from django.shortcuts import render
from django.http import HttpResponse
import openpyxl

# Create your views here.


def index(request):
    return render(request, 'main/index.html')


def kp(request):
    if "GET" == request.method:
        return render(request, 'main/KPE.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.open(excel_file, read_only=True)
        sheet = wb.worksheets[0]

        array = []
        str1 = []
        str2 = []
        str3 = []
        str4 = []
        str5 = []
        str6 = []
        str7 = []
        array1 = []
        str8 = []
        str9 = []
        str10 = []
        str11 = []
        str12 = []
        str13 = []
        str14 = []
        th = []

    for row in range(1, 2):
        th.append([])
        for col in range(2, sheet.max_column):
            th[row-1].append(sheet[row][col].value)

    for row in range(2, 3):
        array.append([])
        for col in range(2, sheet.max_column):
            if col == 2:
                array[row-2].append(sheet[row][col].value)
            else:
                array[row-2].append(int(round((float(sheet[row][col].value))*100)))

    for row in range(3, 4):
        str1.append([])
        for col in range(2, sheet.max_column):
            if col == 2:
                str1[row-3].append(sheet[row][col].value)
            else:
                str1[row-3].append(int(round((float(sheet[row][col].value))*100)))

    for row in range(4, 5):
        str2.append([])
        for col in range(2, sheet.max_column):
            if col == 2:
                str2[row-4].append(sheet[row][col].value)
            else:
                str2[row-4].append(int(round((float(sheet[row][col].value))*100)))

    for row in range(5, 6):
        str3.append([])
        for col in range(2, sheet.max_column):
            if col == 2:
                str3[row-5].append(sheet[row][col].value)
            else:
                str3[row-5].append(int(round((float(sheet[row][col].value))*100)))

    for row in range(6, 7):
        str4.append([])
        for col in range(2, sheet.max_column):
            if col == 2:
                str4[row-6].append(sheet[row][col].value)
            else:
                str4[row-6].append(int(round((float(sheet[row][col].value))*100)))

    for row in range(7, 8):
        str5.append([])
        for col in range(2, sheet.max_column):
            if col == 2:
                str5[row-7].append(sheet[row][col].value)
            else:
                str5[row-7].append(int(round((float(sheet[row][col].value))*100)))

    for row in range(8, 9):
        str6.append([])
        for col in range(2, sheet.max_column):
            if col == 2:
                str6[row-8].append(sheet[row][col].value)
            else:
                str6[row-8].append(int(round((float(sheet[row][col].value))*100)))

    for row in range(9, 10):
        str7.append([])
        for col in range(2, sheet.max_column):
            if col == 2:
                str7[row-9].append(sheet[row][col].value)
            else:
                str7[row-9].append(int(round((float(sheet[row][col].value))*100)))

    for row in range(10, 11):
        array1.append([])
        for col in range(2, sheet.max_column):
            if col == 2:
                array1[row-10].append(sheet[row][col].value)
            else:
                array1[row-10].append(int(round((float(sheet[row][col].value))*100)))

    for row in range(11, 12):
        str8.append([])
        for col in range(2, sheet.max_column):
            if col == 2:
                str8[row-11].append(sheet[row][col].value)
            else:
                str8[row-11].append(int(round((float(sheet[row][col].value))*100)))

    for row in range(12, 13):
        str9.append([])
        for col in range(2, sheet.max_column):
            if col == 2:
                str9[row-12].append(sheet[row][col].value)
            else:
                str9[row-12].append(int(round((float(sheet[row][col].value))*100)))

    for row in range(13, 14):
        str10.append([])
        for col in range(2, sheet.max_column):
            if col == 2:
                str10[row-13].append(sheet[row][col].value)
            else:
                str10[row-13].append(int(round((float(sheet[row][col].value))*100)))

    for row in range(14, 15):
        str11.append([])
        for col in range(2, sheet.max_column):
            if col == 2:
                str11[row-14].append(sheet[row][col].value)
            else:
                str11[row-14].append(int(round((float(sheet[row][col].value))*100)))

    for row in range(15, 16):
        str12.append([])
        for col in range(2, sheet.max_column):
            if col == 2:
                str12[row-15].append(sheet[row][col].value)
            else:
                str12[row-15].append(int(round((float(sheet[row][col].value))*100)))

    for row in range(16, 17):
        str13.append([])
        for col in range(2, sheet.max_column):
            if col == 2:
                str13[row-16].append(sheet[row][col].value)
            else:
                str13[row-16].append(int(round((float(sheet[row][col].value))*100)))

    for row in range(17, 18):
        str14.append([])
        for col in range(2, sheet.max_column):
            if col == 2:
                str14[row-17].append(sheet[row][col].value)
            else:
                str14[row-17].append(int(round((float(sheet[row][col].value))*100)))

    return render(request, 'main/KPE.html', {"array": array, "th": th, "str1": str1, "str2": str2, "str3": str3, "str4":
                                            str4, "str5": str5, "str6": str6, "str7": str7, "str8": str8, "str9": str9,
                                            "str10": str10, "str11": str11, "str12": str12, "str13": str13,
                                            "str14": str14, "array1": array1})


def reg(request):
    return render(request, 'main/Registr.html')


def zav(request):
    return render(request, 'main/Zav.html')


def angcchpp(request):
    return render(request, 'main/ANGCCHPP.html')


def angctf(request):
    return render(request, 'main/ANGCtf.html')


def ano(request):
    return render(request, 'main/ANOCHPP.html')


def azp(request):
    return render(request, 'main/AZPCTS.html')


def kco(request):
    return render(request, 'main/KCo.html')


def kct(request):
    return render(request, 'main/KCt.html')


def kkco(request):
    return render(request, 'main/KKCo.html')


def kkct(request):
    return render(request, 'main/KKCt.html')


def nta(request):
    return render(request, 'main/NTAcds.html')


def ntac(request):
    return render(request, 'main/NTAchpp.html')


def otgr(request):
    return render(request, 'main/OTGRUZKA.html')


def rs(request):
    return render(request, 'main/RScds.html')


def stan1(request):
    return render(request, 'main/STAN1400.html')


def stan2(request):
    return render(request, 'main/STAN2000.html')


def stan3(request):
    return render(request, 'main/STAN2030.html')


def tk(request):
    return render(request, 'main/TKcts.html')


